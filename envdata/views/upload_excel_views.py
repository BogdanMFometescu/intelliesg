from django.db import transaction

from envdata.models import Company, Fuel, NaturalGas, Energy, Sf6, Refrigerant, Travel, Waste
from django.views.generic.edit import FormView
from django.urls import reverse_lazy
from openpyxl import load_workbook
from envdata.forms import ExcelUploadForm
from django.contrib import messages


class ExcelUploadBaseView(FormView):
    form_class = ExcelUploadForm
    template_name = None
    success_url = None
    sheet_name = None
    model_class = None

    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']
        wb = load_workbook(filename=excel_file)

        if self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            messages.error(self.request, 'The specific sheet does not exists')
            return self.form_invalid(form)

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                company_name = row[0]
                if company_name:
                    try:
                        company = Company.objects.get(name=company_name)
                    except Company.DoesNotExist:
                        messages.error(self.request, f"Company '{company_name}' does not exist in the database.")
                        return self.form_invalid(form)

                    profile_id = self.get_profile_id(company)
                    self.create_instance(company, profile_id, row)

        return super().form_valid(form)

    def get_profile_id(self, company):
        raise NotImplementedError("Subclasses of ExcelUploadBaseView must implement get_profile_id method.")

    def create_instance(self, company, profile_id, row):
        raise NotImplementedError("Subclasses of ExcelUploadBaseView must implement create_instance method.")


class ExcelUploadView(ExcelUploadBaseView):
    template_name = 'envdata/upload_excel/upload_fuel_data.html'
    success_url = reverse_lazy('fuel_emissions')
    sheet_name = 'fuel'
    model_class = Fuel

    def get_profile_id(self, company):
        return company.profile.id if hasattr(company, 'profile') else None

    def create_instance(self, company, profile_id, row):
        self.model_class.objects.create(
            company=company,
            profile_id=profile_id,
            month=row[1],
            year=row[2],
            emission_type=row[3],
            emission_scope=row[4],
            fuel_type=row[5],
            fuel_source=row[6],
            fuel_quantity=row[7],
            pollution_norm=row[8],
            activity_type=row[9],
            vehicle_type=row[10],
            measure_unit=row[11],
            emission_factor=row[12]
        )


class ExcelUploadViewForNaturalGas(ExcelUploadBaseView):
    template_name = 'envdata/upload_excel/upload_natural_gas_data.html'
    success_url = reverse_lazy('gas_emissions')
    sheet_name = 'natural_gas'
    model_class = NaturalGas

    def get_profile_id(self, company):
        return company.profile.id if hasattr(company, 'profile') else None

    def create_instance(self, company, profile_id, row):
        self.model_class.objects.create(
            company=company,
            profile_id=profile_id,
            month=row[1],
            year=row[2],
            emission_type=row[3],
            emission_scope=row[4],
            location=row[5],
            gas_quantity=row[6],
            emission_factor=row[7],
            measure_unit=row[8]
        )


class ExcelUploadViewForEnergy(ExcelUploadBaseView):
    template_name = 'envdata/upload_excel/upload_energy_data.html'
    success_url = reverse_lazy('energy_acquisitions')
    sheet_name = 'energy'
    model_class = Energy

    def get_profile_id(self, company):
        return company.profile.id if hasattr(company, 'profile') else None

    def create_instance(self, company, profile_id, row):
        self.model_class.objects.create(
            company=company,
            profile_id=profile_id,
            month = row[1],
            year=row[2],
            emission_type=row[3],
            emission_scope=row[4],
            location=row[5],
            supplier_name=row[6],
            measure_unit=row[7],
            energy_quantity=row[8],
            emission_factor=row[9],
            calculation_method=row[10]
        )


class ExcelUploadViewForSf6(ExcelUploadBaseView):
    template_name = 'envdata/upload_excel/upload_sf6_data.html'
    success_url = reverse_lazy('sf6_emissions')
    sheet_name = 'sf6'
    model_class = Sf6

    def get_profile_id(self, company):
        return company.profile.id if hasattr(company, 'profile') else None

    def create_instance(self, company, profile_id, row):
        self.model_class.objects.create(
            company=company,
            profile_id=profile_id,
            month=row[1],
            year=row[2],
            emission_type=row[3],
            emission_scope=row[4],
            equipment_type=row[5],
            equipment_producer=row[6],
            sf6_quantity=row[7],
            emission_factor=row[8],
        )


class ExcelUploadViewForRefrigerant(ExcelUploadBaseView):
    template_name = 'envdata/upload_excel/upload_sf6_data.html'
    success_url = reverse_lazy('refrigerant_emissions')
    sheet_name = 'refrigerant'
    model_class = Refrigerant

    def get_profile_id(self, company):
        return company.profile.id if hasattr(company, 'profile') else None

    def create_instance(self, company, profile_id, row):
        self.model_class.objects.create(
            company=company,
            profile_id=profile_id,
            month=row[1],
            year=row[2],
            emission_type=row[3],
            emission_scope=row[4],
            refrigerant_type=row[5],
            refrigerant_quantity=row[6],
            emission_factor=row[7],

        )


class ExcelUploadViewForTravel(ExcelUploadBaseView):
    template_name = 'envdata/upload_excel/upload_travel_data.html'
    success_url = reverse_lazy('travels')
    sheet_name = 'travel'
    model_class = Travel

    def get_profile_id(self, company):
        return company.profile.id if hasattr(company, 'profile') else None

    def create_instance(self, company, profile_id, row):
        self.model_class.objects.create(
            company=company,
            profile_id=profile_id,
            month=row[1],
            year=row[2],
            emission_type=row[3],
            emission_scope=row[4],
            origin=row[5],
            destination=row[6],
            emission_factor=row[7],
            distance=row[8],
            fuel_consumption=row[9],
        )


class ExcelUploadViewForWaste(ExcelUploadBaseView):
    template_name = 'envdata/upload_excel/upload_waste_data.html'
    success_url = reverse_lazy('wastes')
    sheet_name = 'waste'
    model_class = Waste

    def get_profile_id(self, company):
        return company.profile.id if hasattr(company, 'profile') else None

    def create_instance(self, company, profile_id, row):
        self.model_class.objects.create(
            company=company,
            profile_id=profile_id,
            month=row[1],
            year=row[2],
            emission_type=row[3],
            emission_scope=row[4],
            waste_name=row[5],
            quantity_recycled=row[6],
            quantity_disposed=row[7],
            quantity_land_filled=row[8],
            emission_factor=row[9],
        )
